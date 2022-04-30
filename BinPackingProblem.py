import openpyxl
import numpy as np

class Bin(object):

    def __init__(self):
        self.items = []  
        self.total_weight = 0  

    def add_item(self, item): 
        self.items.append(item)
        self.total_weight += item


def read(f):

    book = openpyxl.load_workbook(f)  
    sheet = book.active      
    items_list =[]
    n = sheet.cell(row = 2, column = 1).value
    c = sheet.cell(row = 2, column = 2).value
    for i in range(2,n+2):  
        cell_obj = sheet.cell(row = i, column = 3)
        items_list.append(cell_obj.value) 
    return items_list,c  


def ffa(items_list, bin_capacity):
    
    bins =[]
    randomised_np_list = np.random.permutation(items_list) 
    items_list = randomised_np_list.tolist() 
    
    for item in items_list:

        for bin in bins:
            if bin.total_weight + item <= bin_capacity: 
                bin.add_item(item)  
                break
        else:
            
            bin = Bin()
            bin.add_item(item)
            bins.append(bin)

    return bins

def ffda(items_list, bin_capacity):

    decreased_list = sorted(items_list,reverse=True) 
    bins =[]
    for item in decreased_list:

        for bin in bins:
            if bin.total_weight + item <= bin_capacity: 
                bin.add_item(item)  
                break
        else:
            
            bin = Bin()
            bin.add_item(item)
            bins.append(bin)

    return bins
    
if __name__ == '__main__': 
          
    result=read("Excel_A.xlsx")
    items_list,c = result[0],result[1] 
    bins_ffa = ffa(items_list,c)
    bins_ffda = ffda(items_list,c)
     
    print("ffa solution is using: ",len(bins_ffa), "bins")
    print("ffda solution is using: ",len(bins_ffda), "bins")

    print(list(bins_ffa))
    print(list(bins_ffda))
    
